"""Builds the BAR RADAR insights workbook from Phase 3's structured data
(menu_items, brand_mentions) - top brands, category share, pricing,
alcoholic/non-alcoholic split.

This is a fully regenerated report (rebuilt from scratch every run, not a
hand-edited model), so cells hold computed values rather than live Excel
formulas - that avoids needing LibreOffice on the CI runner just to
recalculate a report nobody edits by hand.

Usage:
    python -m src.pipeline.generate_report
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from db.database import (
    alcoholic_split,
    brand_leaderboard,
    category_share,
    export_menu_items_rows,
    spirit_category_share,
)
from db.migrate import run as run_migrations
from src.utils.config import REPO_ROOT
from src.utils.logging_utils import get_logger

log = get_logger(__name__)

FONT = "Arial"
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
HEADER_FILL = "4472C4"
TITLE_FONT = Font(name=FONT, bold=True, size=14)
BODY_FONT = Font(name=FONT)


def _style_header_row(ws, row: int, num_cols: int) -> None:
    from openpyxl.styles import PatternFill
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")


def _autosize_columns(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_table(ws, headers: list[str], rows: list[list], start_row: int = 1) -> int:
    for col, header in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col, value=header)
    _style_header_row(ws, start_row, len(headers))
    for r, row in enumerate(rows, start=start_row + 1):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = BODY_FONT
    return start_row + len(rows) + 1


def build_workbook(output_path) -> None:
    items = [dict(r) for r in export_menu_items_rows()]
    brands = [dict(r) for r in brand_leaderboard()]
    categories = [dict(r) for r in category_share()]
    spirits = [dict(r) for r in spirit_category_share()]
    alcoholic = [dict(r) for r in alcoholic_split()]

    wb = Workbook()

    # --- Overview -----------------------------------------------------
    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = "BAR RADAR - Menu Insights"
    ws["A1"].font = TITLE_FONT
    venues = {i["venue_id"] for i in items}
    priced_items = [i for i in items if i.get("price_value") is not None]
    avg_price = round(sum(i["price_value"] for i in priced_items) / len(priced_items), 2) if priced_items else None

    overview_rows = [
        ["Venues with analyzed menu items", len(venues)],
        ["Total menu items extracted", len(items)],
        ["Items with a recognized price", len(priced_items)],
        ["Average item price (EUR, all categories)", avg_price],
        ["Distinct brands mentioned", len({b["brand_name"] for b in brands})],
        ["Total brand mentions", sum(b["total_mentions"] for b in brands)],
    ]
    for r, (label, value) in enumerate(overview_rows, start=3):
        ws.cell(row=r, column=1, value=label).font = BODY_FONT
        ws.cell(row=r, column=2, value=value).font = BODY_FONT
    _autosize_columns(ws, {1: 40, 2: 16})

    note_row = 3 + len(overview_rows) + 1
    ws.cell(row=note_row, column=1,
            value="Note: parsing is best-effort against real, messy PDF/HTML text - see "
                  "parse_confidence in 'Raw Menu Items' for how much to trust each row.").font = \
        Font(name=FONT, italic=True, size=9, color="808080")

    # --- Top Brands -----------------------------------------------------
    ws = wb.create_sheet("Top Brands")
    rows = [[b["brand_name"], b["spirit_category"], b["total_mentions"], b["venue_count"]] for b in brands]
    _write_table(ws, ["Brand", "Spirit Category", "Total Mentions", "Venues Calling It"], rows)
    _autosize_columns(ws, {1: 22, 2: 18, 3: 16, 4: 18})

    # --- Beverage Category Share -----------------------------------------
    ws = wb.create_sheet("Category Share")
    rows = [
        [c["beverage_category"], c["item_count"],
         round(c["avg_price"], 2) if c["avg_price"] is not None else None,
         c["min_price"], c["max_price"]]
        for c in categories
    ]
    _write_table(ws, ["Beverage Category", "Item Count", "Avg Price (EUR)", "Min Price", "Max Price"], rows)
    _autosize_columns(ws, {1: 20, 2: 14, 3: 16, 4: 12, 5: 12})

    # --- Spirit Category Share -----------------------------------------
    ws = wb.create_sheet("Spirit Category Share")
    rows = [
        [s["spirit_category"], s["item_count"], round(s["avg_price"], 2) if s["avg_price"] is not None else None]
        for s in spirits
    ]
    _write_table(ws, ["Spirit Category", "Item Count", "Avg Price (EUR)"], rows)
    _autosize_columns(ws, {1: 20, 2: 14, 3: 16})

    # --- Alcoholic vs Non-Alcoholic -----------------------------------------
    ws = wb.create_sheet("Alcoholic Split")
    rows = [[a["label"], a["item_count"]] for a in alcoholic]
    _write_table(ws, ["Type", "Item Count"], rows)
    _autosize_columns(ws, {1: 20, 2: 14})

    # --- Raw Menu Items (reference) -----------------------------------------
    ws = wb.create_sheet("Raw Menu Items")
    rows = [
        [i["venue_name"], i["city"], i["tier"], i["item_name"], i["price_value"], i["price_currency"],
         i["beverage_category"], i["spirit_category"],
         ("Yes" if i["is_alcoholic"] == 1 else "No" if i["is_alcoholic"] == 0 else "Unknown"),
         i["ingredients_text"], round(i["parse_confidence"], 2) if i["parse_confidence"] is not None else None,
         i["parse_method"]]
        for i in items
    ]
    _write_table(ws, [
        "Venue", "City", "Tier", "Item Name", "Price", "Currency", "Beverage Category",
        "Spirit Category", "Alcoholic", "Ingredients", "Parse Confidence", "Parse Method",
    ], rows)
    _autosize_columns(ws, {1: 22, 2: 14, 3: 8, 4: 26, 5: 10, 6: 10, 7: 18, 8: 16, 9: 12, 10: 40, 11: 16, 12: 26})

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info("Wrote insights workbook to %s (%d items, %d brands)", output_path, len(items), len(brands))


def run() -> None:
    run_migrations()
    output_path = REPO_ROOT / "data" / "exports" / "bar_radar_insights.xlsx"
    build_workbook(output_path)


if __name__ == "__main__":
    run()
